import io
import re
import openpyxl
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from ..core.logging import get_logger
from ..core.models import GmpFactory

logger = get_logger("scraper.gmp_parser")


def _clean_str(val: any) -> str | None:
    """Clean and standardize strings from Excel cells."""
    if val is None:
        return None
    val_str = str(val).strip()
    if not val_str:
        return None
    # Thay thế nhiều khoảng trắng hoặc newline liên tiếp bằng 1 khoảng trắng
    return re.sub(r"\s+", " ", val_str)


def parse_gmp_sheet(file_bytes: bytes) -> list[dict]:
    """Parse the WHO-GMP / EU-GMP manufacturing or packaging plants Excel sheet (cn92 list 1).
    
    Dynamically maps columns based on the header row containing 'TÊN CƠ SỞ'.
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
    sheet = wb.active
    
    # 1. Tìm dòng tiêu đề chứa "TÊN CƠ SỞ"
    header_row_idx = None
    header_cells = None
    for row_idx, row in enumerate(sheet.iter_rows(values_only=True), 1):
        row_strs = [str(x).strip().upper() for x in row if x is not None]
        if "TÊN CƠ SỞ" in row_strs:
            header_row_idx = row_idx
            header_cells = [str(x).strip() if x is not None else "" for x in row]
            break
            
    if header_row_idx is None:
        logger.warning("Could not find header row containing 'TÊN CƠ SỞ'. Using defaults.")
        name_idx = 2
        address_idx = 4
        scope_idx = 5
        standard_idx = 7
        authority_idx = 8
    else:
        logger.info(f"Found header row at index {header_row_idx}")
        name_idx = None
        address_idx = None
        scope_idx = None
        standard_idx = None
        authority_idx = None
        
        for idx, cell in enumerate(header_cells):
            cell_upper = cell.upper()
            if "TÊN CƠ SỞ" in cell_upper:
                name_idx = idx
            elif "ĐỊA CHỈ CƠ SỞ" in cell_upper:
                address_idx = idx
            elif "PHẠM VI" in cell_upper:
                scope_idx = idx
            elif "TIÊU CHUẨN" in cell_upper:
                standard_idx = idx
            elif "CƠ QUAN" in cell_upper:
                authority_idx = idx
                
        # Fallback to defaults if any column is missing
        if name_idx is None: name_idx = 2
        if address_idx is None: address_idx = 4
        if scope_idx is None: scope_idx = 5
        if standard_idx is None: standard_idx = 7
        if authority_idx is None: authority_idx = 8
        
    parsed_rows = []
    row_count = 0
    start_row = (header_row_idx if header_row_idx else 6) + 1
    
    for row_idx, row in enumerate(sheet.iter_rows(values_only=True), 1):
        if row_idx < start_row:
            continue
            
        if len(row) <= max(name_idx, address_idx, scope_idx, standard_idx, authority_idx):
            continue
            
        tt_val = _clean_str(row[0])
        # Kiểm tra xem dòng này có phải dòng dữ liệu thực tế (bắt đầu bằng số thứ tự)
        if not tt_val or not tt_val.isdigit():
            continue
            
        factory_name = _clean_str(row[name_idx])
        address = _clean_str(row[address_idx])
        
        if not factory_name or not address:
            continue
            
        standard_val = _clean_str(row[standard_idx])
        # Safeguard: if standard value looks like a date/timestamp, filter it out
        if standard_val and (re.search(r"\d{4}-\d{2}-\d{2}", standard_val) or "00:00:00" in standard_val):
            logger.warning(f"Discarding date-like standard value: '{standard_val}' for factory '{factory_name}'")
            standard_val = None

        parsed_rows.append({
            "factory_name": factory_name,
            "address": address,
            "scope": _clean_str(row[scope_idx]),
            "standard": standard_val,
            "authority": _clean_str(row[authority_idx]),
            "headquarters_address": None,
            "location_name": None,
            "responsible_pharmacist": None,
            "certificate_license": None
        })
        row_count += 1
        
    logger.info(f"Parsed {row_count} GMP manufacturing facilities from Excel sheet.")
    return parsed_rows


def parse_license_sheet(file_bytes: bytes) -> list[dict]:
    """Parse the DKKD Dược (Drug Business Certificate) Excel sheet (cn92 list 2).
    
    Header row: row 4. Data starts at row 7.
    Indices:
      - Col 0 (TT): Index number (must be numeric)
      - Col 1 (Tên cơ sở kinh doanh)
      - Col 2 (Địa chỉ trụ sở chính)
      - Col 3 (Tên địa điểm kinh doanh)
      - Col 4 (Địa điểm kinh doanh)
      - Col 5 (Người chịu trách nhiệm chuyên môn)
      - Col 6 (Phạm vi hoạt động sản xuất thuốc)
      - Col 7 (Giấy CN đủ điều kiện kinh doanh dược)
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
    sheet = wb.active
    
    parsed_rows = []
    row_count = 0
    
    for row_idx, row in enumerate(sheet.iter_rows(values_only=True), 1):
        if row_idx < 5:  # Bỏ qua dòng tiêu đề ban đầu
            continue
            
        if len(row) < 8:
            continue
            
        tt_val = _clean_str(row[0])
        if not tt_val or not tt_val.isdigit():
            continue
            
        factory_name = _clean_str(row[1])
        address = _clean_str(row[4])  # Sử dụng Địa điểm kinh doanh làm địa chỉ sản xuất chính
        
        if not factory_name or not address:
            continue
            
        parsed_rows.append({
            "factory_name": factory_name,
            "address": address,
            "scope": _clean_str(row[6]),
            "standard": "GMP",  # Giá trị tiêu chuẩn mặc định cho cơ sở sản xuất có giấy phép ĐKKD
            "authority": "Bộ Y tế Việt Nam",
            "headquarters_address": _clean_str(row[2]),
            "location_name": _clean_str(row[3]),
            "responsible_pharmacist": _clean_str(row[5]),
            "certificate_license": _clean_str(row[7])
        })
        row_count += 1
        
    logger.info(f"Parsed {row_count} GMP licensed business facilities from Excel sheet.")
    return parsed_rows


async def sync_gmp_data(session: AsyncSession, category: str, parsed_rows: list[dict]) -> list[dict]:
    """Sync Excel/PDF rows with the DB and return newly added or updated factories."""
    newly_added = []
    
    for data in parsed_rows:
        # Tìm kiếm cơ sở xem đã tồn tại trong database chưa (theo category, tên cơ sở và địa chỉ)
        stmt = select(GmpFactory).where(
            GmpFactory.category == category,
            GmpFactory.factory_name == data["factory_name"],
            GmpFactory.address == data["address"]
        )
        result = await session.execute(stmt)
        existing = result.scalar_one_or_none()
        
        if existing is None:
            # Phát hiện nhà máy hoàn toàn mới!
            factory = GmpFactory(
                category=category,
                factory_name=data["factory_name"],
                address=data["address"],
                scope=data["scope"],
                standard=data["standard"],
                authority=data["authority"],
                headquarters_address=data["headquarters_address"],
                location_name=data["location_name"],
                responsible_pharmacist=data["responsible_pharmacist"],
                certificate_license=data["certificate_license"]
            )
            session.add(factory)
            
            # Copy data and mark as new
            item_data = dict(data)
            item_data["_change_type"] = "new"
            newly_added.append(item_data)
        else:
            # Cơ sở đã tồn tại -> Kiểm tra xem có bất kỳ trường thông tin nào thay đổi (vd cập nhật phạm vi, gia hạn)
            changed = False
            changes = {}
            for field in [
                "scope", "standard", "authority", "headquarters_address", 
                "location_name", "responsible_pharmacist", "certificate_license"
            ]:
                new_val = data.get(field)
                old_val = getattr(existing, field)
                if new_val is not None and old_val != new_val:
                    changes[field] = (old_val, new_val)
                    setattr(existing, field, new_val)
                    changed = True
            if changed:
                session.add(existing)
                
                # Copy data and mark as updated with changes details
                item_data = dict(data)
                item_data["_change_type"] = "update"
                item_data["_changes"] = changes
                newly_added.append(item_data)
                
    await session.flush()
    return newly_added
